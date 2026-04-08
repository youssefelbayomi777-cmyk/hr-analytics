import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference, LineChart
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
import json

# Load analysis results
with open('/home/z/my-project/download/analysis_results.json', 'r') as f:
    results = json.load(f)

# Load original data
df = pd.read_excel('/home/z/my-project/upload/01JTD2XV9ZPTCK9QDKE0TG3R87.xlsx')

# Add computed columns
df.columns = ['Employee_ID', 'Name', 'Department', 'Date_of_Hire', 
              'Monthly_Absence_Rate', 'Overtime_Hours', 
              'Performance_Rating', 'Turnover_3Y', 'Satisfaction']
df['Date_of_Hire'] = pd.to_datetime(df['Date_of_Hire'])
df['Tenure_Years'] = ((pd.Timestamp.now() - df['Date_of_Hire']).dt.days / 365.25).round(1)
df['Turnover_Flag'] = (df['Turnover_3Y'] > 0).astype(int)
df['Date_of_Hire'] = df['Date_of_Hire'].dt.strftime('%Y-%m-%d')

wb = Workbook()

# ============================================================
# Sheet 1: Cleaned Data
# ============================================================
ws1 = wb.active
ws1.title = "Cleaned Data"
ws1.sheet_properties.tabColor = "27AE60"

# Headers
headers = ['Employee ID', 'Name', 'Department', 'Date of Hire', 
           'Monthly Absence Rate (%)', 'Overtime Hours', 
           'Performance Rating (1-5)', 'Turnover (3 Years)', 
           'Satisfaction (1-10)', 'Tenure (Years)', 'Turnover Flag']
header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Data
for row_idx, row in df.iterrows():
    for col_idx, col in enumerate(headers):
        if col == 'Employee ID':
            val = row['Employee_ID']
        elif col == 'Name':
            val = row['Name']
        elif col == 'Department':
            val = row['Department']
        elif col == 'Date of Hire':
            val = row['Date_of_Hire']
        elif col == 'Monthly Absence Rate (%)':
            val = row['Monthly_Absence_Rate']
        elif col == 'Overtime Hours':
            val = row['Overtime_Hours']
        elif col == 'Performance Rating (1-5)':
            val = row['Performance_Rating']
        elif col == 'Turnover (3 Years)':
            val = row['Turnover_3Y']
        elif col == 'Satisfaction (1-10)':
            val = row['Satisfaction']
        elif col == 'Tenure (Years)':
            val = row['Tenure_Years']
        elif col == 'Turnover Flag':
            val = row['Turnover_Flag']
        
        cell = ws1.cell(row=row_idx + 2, column=col_idx + 1, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        if col_idx >= 4 and col_idx <= 9:
            cell.number_format = '0.0'

# Auto-width
for col in range(1, len(headers) + 1):
    ws1.column_dimensions[get_column_letter(col)].width = 20

ws1.auto_filter.ref = f"A1:K{len(df) + 1}"

# ============================================================
# Sheet 2: Descriptive Statistics
# ============================================================
ws2 = wb.create_sheet("Descriptive Statistics")
ws2.sheet_properties.tabColor = "3498DB"

stats_headers = ['Metric', 'Mean', 'Median', 'Std Dev', 'Min', 'Max', 'Q1', 'Q3']
metric_names = {
    'Monthly_Absence_Rate': 'Monthly Absence Rate (%)',
    'Overtime_Hours': 'Overtime Hours',
    'Performance_Rating': 'Performance Rating (1-5)',
    'Satisfaction': 'Satisfaction Score (1-10)',
    'Turnover_3Y': 'Turnover Count (3Y)',
    'Tenure_Years': 'Tenure (Years)'
}

for col, h in enumerate(stats_headers, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

row_idx = 2
for metric_key, metric_label in metric_names.items():
    s = results['descriptive_stats'].get(metric_key, {})
    values = [metric_label, s.get('mean', ''), s.get('median', ''), s.get('std', ''),
              s.get('min', ''), s.get('max', ''), s.get('q1', ''), s.get('q3', '')]
    for col, val in enumerate(values, 1):
        cell = ws2.cell(row=row_idx, column=col, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        if col >= 2:
            cell.number_format = '0.00'
    row_idx += 1

for col in range(1, len(stats_headers) + 1):
    ws2.column_dimensions[get_column_letter(col)].width = 22

# ============================================================
# Sheet 3: Department Analysis
# ============================================================
ws3 = wb.create_sheet("Department Analysis")
ws3.sheet_properties.tabColor = "E74C3C"

dept_headers = ['Department', 'Employees', 'Avg Absence (%)', 'Avg Overtime (h)',
                'Avg Performance', 'Avg Satisfaction', 'Avg Turnover', 'Avg Tenure (Yrs)']

for col, h in enumerate(dept_headers, 1):
    cell = ws3.cell(row=1, column=col, value=h)
    cell.fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = thin_border

row_idx = 2
dept_stats = results['department_stats']
for dept in sorted(dept_stats.keys()):
    d = dept_stats[dept]
    values = [dept, d.get('Count', 0), d.get('Avg_Absence', ''),
              d.get('Avg_Overtime', ''), d.get('Avg_Performance', ''),
              d.get('Avg_Satisfaction', ''), d.get('Avg_Turnover', ''),
              d.get('Avg_Tenure', '')]
    for col, val in enumerate(values, 1):
        cell = ws3.cell(row=row_idx, column=col, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        if col >= 3:
            cell.number_format = '0.00'
    row_idx += 1

for col in range(1, len(dept_headers) + 1):
    ws3.column_dimensions[get_column_letter(col)].width = 18

# ============================================================
# Sheet 4: KPIs
# ============================================================
ws4 = wb.create_sheet("KPIs")
ws4.sheet_properties.tabColor = "F39C12"

kpis = results['kpis']
kpi_items = [
    ('Overall Absence Rate (%)', kpis['overall_absence_rate']),
    ('Average Performance Rating', kpis['avg_performance']),
    ('Average Satisfaction Score', kpis['avg_satisfaction']),
    ('Turnover Rate - 3 Years (%)', kpis['turnover_rate_3y']),
    ('Average Overtime Hours', kpis['avg_overtime']),
    ('Average Employee Tenure (Years)', kpis['avg_tenure']),
    ('Overtime-Satisfaction Correlation', kpis['ot_satisfaction_correlation']),
    ('Total Employees', results['data_overview']['total_employees']),
    ('Number of Departments', results['data_overview']['departments']),
]

kpi_header_fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
for col, h in enumerate(['KPI', 'Value'], 1):
    cell = ws4.cell(row=1, column=col, value=h)
    cell.fill = kpi_header_fill
    cell.font = Font(color="FFFFFF", bold=True, size=12)
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

for row_idx, (label, value) in enumerate(kpi_items, 2):
    cell_l = ws4.cell(row=row_idx, column=1, value=label)
    cell_l.border = thin_border
    cell_l.font = Font(bold=True)
    cell_v = ws4.cell(row=row_idx, column=2, value=value)
    cell_v.border = thin_border
    cell_v.alignment = Alignment(horizontal='center')
    cell_v.number_format = '0.00'

ws4.column_dimensions['A'].width = 35
ws4.column_dimensions['B'].width = 15

# ============================================================
# Sheet 5: Predictive Analysis
# ============================================================
ws5 = wb.create_sheet("Predictive Analysis")
ws5.sheet_properties.tabColor = "9B59B6"

pred = results['predictive']
pred_headers = ['Model / Feature', 'Value']

for col, h in enumerate(pred_headers, 1):
    cell = ws5.cell(row=1, column=col, value=h)
    cell.fill = PatternFill(start_color="8E44AD", end_color="8E44AD", fill_type="solid")
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

pred_items = [
    ('Logistic Regression Accuracy', pred['logistic_regression_accuracy']),
    ('Decision Tree Accuracy', pred['decision_tree_accuracy']),
    ('', ''),
    ('Feature Importance (Decision Tree)', ''),
]
for feat, imp in pred['feature_importance'].items():
    pred_items.append((f'  {feat}', imp))

pred_items.append(('', ''))
pred_items.append(('Logistic Regression Coefficients', ''))
for feat, coef in pred['lr_coefficients'].items():
    pred_items.append((f'  {feat}', coef))

for row_idx, (label, value) in enumerate(pred_items, 2):
    cell_l = ws5.cell(row=row_idx, column=1, value=label)
    cell_l.border = thin_border
    cell_v = ws5.cell(row=row_idx, column=2, value=value)
    cell_v.border = thin_border
    cell_v.alignment = Alignment(horizontal='center')
    if isinstance(value, float):
        cell_v.number_format = '0.0000'
    if label and not label.startswith('  '):
        cell_l.font = Font(bold=True)

ws5.column_dimensions['A'].width = 40
ws5.column_dimensions['B'].width = 15

# ============================================================
# Add Charts to Department Analysis sheet
# ============================================================

# Bar chart: Absence by Department
chart1 = BarChart()
chart1.type = "col"
chart1.title = "Average Absence Rate by Department"
chart1.y_axis.title = "Absence Rate (%)"
chart1.x_axis.title = "Department"
chart1.style = 10
data_ref = Reference(ws3, min_col=3, min_row=1, max_row=len(dept_stats) + 1)
cats_ref = Reference(ws3, min_col=1, min_row=2, max_row=len(dept_stats) + 1)
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats_ref)
chart1.shape = 4
chart1.width = 20
chart1.height = 12
ws3.add_chart(chart1, "A12")

# Bar chart: Performance by Department
chart2 = BarChart()
chart2.type = "col"
chart2.title = "Average Performance by Department"
chart2.y_axis.title = "Performance (1-5)"
chart2.style = 10
data_ref2 = Reference(ws3, min_col=5, min_row=1, max_row=len(dept_stats) + 1)
chart2.add_data(data_ref2, titles_from_data=True)
chart2.set_categories(cats_ref)
chart2.width = 20
chart2.height = 12
ws3.add_chart(chart2, "A28")

# Save
output_path = '/home/z/my-project/download/HR_Analysis_Report.xlsx'
wb.save(output_path)
print(f"✓ Excel file saved: {output_path}")

